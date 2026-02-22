"""Generate the Elron Simulation Data Excel file with 5 tabs."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# Styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_sheet(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.freeze_panes = "A2"

def add_row(ws, row_num, data):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

# =========================================================================
# TAB 1: Tenants
# =========================================================================
ws_tenants = wb.active
ws_tenants.title = "Tenants"
headers = ["First Name", "Last Name", "Email", "Phone Number"]
style_sheet(ws_tenants, headers)

tenants = [
    # Single-unit
    ("James", "Smith", "james.smith.tenant@gmail.com", "(555) 201-1001"),
    ("Maria", "Garcia", "maria.garcia.apt4b@gmail.com", "(555) 201-1002"),
    ("David", "Chen", "david.chen.renter@gmail.com", "(555) 201-1003"),
    ("Sarah", "Johnson", "sarah.j.tenant@gmail.com", "(555) 201-1004"),
    ("Michael", "Brown", "mbrown.tenant@gmail.com", "(555) 201-1005"),
    ("Emily", "Davis", "emily.davis.lease@gmail.com", "(555) 201-1006"),
    ("Robert", "Wilson", "rwilson.apt@gmail.com", "(555) 201-1007"),
    ("Jessica", "Martinez", "jessica.m.renter@gmail.com", "(555) 201-1008"),
    ("William", "Anderson", "william.a.tenant@gmail.com", "(555) 201-1009"),
    ("Linda", "Taylor", "linda.taylor.apt@gmail.com", "(555) 201-1010"),
    # 45 Maple Avenue
    ("Christopher", "Thomas", "c.thomas.renter@gmail.com", "(555) 301-1001"),
    ("Patricia", "Jackson", "p.jackson.tenant@gmail.com", "(555) 301-1002"),
    ("Daniel", "White", "d.white.apt12@gmail.com", "(555) 301-1003"),
    ("Barbara", "Harris", "barbara.h.lease@gmail.com", "(555) 301-1004"),
    # 78 Pine Road
    ("Matthew", "Martin", "matt.martin.tenant@gmail.com", "(555) 302-1001"),
    ("Ashley", "Robinson", "ashley.robinson.unit@gmail.com", "(555) 302-1002"),
    ("Kevin", "Lee", "kevin.lee.renter@gmail.com", "(555) 302-1003"),
    ("Stephanie", "Hall", "s.hall.tenant@gmail.com", "(555) 302-1004"),
    # 231 Elm Blvd
    ("Brian", "Walker", "brian.w.apt@gmail.com", "(555) 303-1001"),
    ("Amanda", "Allen", "a.allen.lease@gmail.com", "(555) 303-1002"),
    ("Ryan", "Young", "ryan.young.tenant@gmail.com", "(555) 303-1003"),
    ("Nicole", "King", "nicole.king.apt@gmail.com", "(555) 303-1004"),
    # 12 Oak Street
    ("Joshua", "Wright", "j.wright.renter@gmail.com", "(555) 304-1001"),
    ("Megan", "Scott", "megan.scott.unit@gmail.com", "(555) 304-1002"),
    ("Tyler", "Green", "tyler.green.apt@gmail.com", "(555) 304-1003"),
    ("Kayla", "Adams", "kayla.adams.tenant@gmail.com", "(555) 304-1004"),
    # 103 Birch Court
    ("Nathan", "Baker", "n.baker.renter@gmail.com", "(555) 305-1001"),
    ("Brittany", "Carter", "b.carter.lease@gmail.com", "(555) 305-1002"),
    ("Andrew", "Mitchell", "andrew.m.tenant@gmail.com", "(555) 305-1003"),
    ("Samantha", "Perez", "s.perez.apt@gmail.com", "(555) 305-1004"),
    # 55 Walnut Drive
    ("Justin", "Roberts", "justin.r.renter@gmail.com", "(555) 306-1001"),
    ("Lauren", "Turner", "l.turner.tenant@gmail.com", "(555) 306-1002"),
    ("Brandon", "Phillips", "b.phillips.apt@gmail.com", "(555) 306-1003"),
    ("Heather", "Campbell", "h.campbell.lease@gmail.com", "(555) 306-1004"),
    # 18 Spruce Way
    ("Eric", "Parker", "eric.parker.unit@gmail.com", "(555) 307-1001"),
    ("Amber", "Evans", "amber.evans.tenant@gmail.com", "(555) 307-1002"),
    ("Jonathan", "Edwards", "jon.edwards.apt@gmail.com", "(555) 307-1003"),
    ("Rachel", "Collins", "rachel.c.renter@gmail.com", "(555) 307-1004"),
    # 400 Willow Terrace
    ("Aaron", "Stewart", "aaron.s.tenant@gmail.com", "(555) 308-1001"),
    ("Danielle", "Sanchez", "d.sanchez.apt@gmail.com", "(555) 308-1002"),
    ("Zachary", "Morris", "z.morris.renter@gmail.com", "(555) 308-1003"),
    ("Melissa", "Rogers", "melissa.r.tenant@gmail.com", "(555) 308-1004"),
    # 7 Ash Place
    ("Patrick", "Reed", "p.reed.apt@gmail.com", "(555) 309-1001"),
    ("Tiffany", "Cook", "tiffany.cook.lease@gmail.com", "(555) 309-1002"),
    ("Gregory", "Morgan", "g.morgan.tenant@gmail.com", "(555) 309-1003"),
    ("Crystal", "Bell", "crystal.bell.apt@gmail.com", "(555) 309-1004"),
    # 92 Hawthorn Gardens
    ("Sean", "Murphy", "sean.murphy.renter@gmail.com", "(555) 310-1001"),
    ("Vanessa", "Rivera", "v.rivera.tenant@gmail.com", "(555) 310-1002"),
    ("Derek", "Cooper", "derek.cooper.apt@gmail.com", "(555) 310-1003"),
    ("Monica", "Richardson", "m.richardson.lease@gmail.com", "(555) 310-1004"),
]

for i, t in enumerate(tenants, 2):
    add_row(ws_tenants, i, t)
auto_width(ws_tenants)

# =========================================================================
# TAB 2: Landlord
# =========================================================================
ws_landlord = wb.create_sheet("Landlord")
headers = ["First Name", "Last Name", "Email", "Phone Number"]
style_sheet(ws_landlord, headers)
add_row(ws_landlord, 2, ("Elron", "Properties", "elron.properties.mgmt@gmail.com", "(555) 100-0001"))
add_row(ws_landlord, 3, ("Ben", "Young", "ben@elron.ai", "(555) 100-0002"))
auto_width(ws_landlord)

# =========================================================================
# TAB 3: Properties
# =========================================================================
ws_props = wb.create_sheet("Properties")
headers = ["Property ID", "Name", "Address"]
style_sheet(ws_props, headers)

properties = [
    ("PROP-001", "14 Birchwood Lane", "14 Birchwood Lane"),
    ("PROP-002", "88 Sunset Drive", "88 Sunset Drive"),
    ("PROP-003", "203 Elm Street", "203 Elm Street"),
    ("PROP-004", "31 Harbor View Road", "31 Harbor View Road"),
    ("PROP-005", "156 Magnolia Court", "156 Magnolia Court"),
    ("PROP-006", "72 Ridgeline Terrace", "72 Ridgeline Terrace"),
    ("PROP-007", "445 Lakeshore Blvd", "445 Lakeshore Blvd"),
    ("PROP-008", "19 Foxglove Way", "19 Foxglove Way"),
    ("PROP-009", "267 Coventry Circle", "267 Coventry Circle"),
    ("PROP-010", "8 Whispering Pines Road", "8 Whispering Pines Road"),
    ("PROP-011", "45 Maple Avenue", "45 Maple Avenue"),
    ("PROP-012", "78 Pine Road", "78 Pine Road"),
    ("PROP-013", "231 Elm Blvd", "231 Elm Blvd"),
    ("PROP-014", "12 Oak Street", "12 Oak Street"),
    ("PROP-015", "103 Birch Court", "103 Birch Court"),
    ("PROP-016", "55 Walnut Drive", "55 Walnut Drive"),
    ("PROP-017", "18 Spruce Way", "18 Spruce Way"),
    ("PROP-018", "400 Willow Terrace", "400 Willow Terrace"),
    ("PROP-019", "7 Ash Place", "7 Ash Place"),
    ("PROP-020", "92 Hawthorn Gardens", "92 Hawthorn Gardens"),
]

for i, p in enumerate(properties, 2):
    add_row(ws_props, i, p)
auto_width(ws_props)

# =========================================================================
# TAB 4: Units
# =========================================================================
ws_units = wb.create_sheet("Units")
headers = ["Unit Number", "Property Address", "Building Number", "Tenant Email", "Landlord Email"]
style_sheet(ws_units, headers)

landlord_email = "elron.properties.mgmt@gmail.com"

units = [
    # Single-unit properties (unit = "—")
    ("—", "14 Birchwood Lane", "PROP-001", "james.smith.tenant@gmail.com", landlord_email),
    ("—", "88 Sunset Drive", "PROP-002", "maria.garcia.apt4b@gmail.com", landlord_email),
    ("—", "203 Elm Street", "PROP-003", "david.chen.renter@gmail.com", landlord_email),
    ("—", "31 Harbor View Road", "PROP-004", "sarah.j.tenant@gmail.com", landlord_email),
    ("—", "156 Magnolia Court", "PROP-005", "mbrown.tenant@gmail.com", landlord_email),
    ("—", "72 Ridgeline Terrace", "PROP-006", "emily.davis.lease@gmail.com", landlord_email),
    ("—", "445 Lakeshore Blvd", "PROP-007", "rwilson.apt@gmail.com", landlord_email),
    ("—", "19 Foxglove Way", "PROP-008", "jessica.m.renter@gmail.com", landlord_email),
    ("—", "267 Coventry Circle", "PROP-009", "william.a.tenant@gmail.com", landlord_email),
    ("—", "8 Whispering Pines Road", "PROP-010", "linda.taylor.apt@gmail.com", landlord_email),
    # 45 Maple Avenue
    ("1A", "45 Maple Avenue", "PROP-011", "c.thomas.renter@gmail.com", landlord_email),
    ("2B", "45 Maple Avenue", "PROP-011", "p.jackson.tenant@gmail.com", landlord_email),
    ("3C", "45 Maple Avenue", "PROP-011", "d.white.apt12@gmail.com", landlord_email),
    ("4D", "45 Maple Avenue", "PROP-011", "barbara.h.lease@gmail.com", landlord_email),
    # 78 Pine Road
    ("101", "78 Pine Road", "PROP-012", "matt.martin.tenant@gmail.com", landlord_email),
    ("202", "78 Pine Road", "PROP-012", "ashley.robinson.unit@gmail.com", landlord_email),
    ("303", "78 Pine Road", "PROP-012", "kevin.lee.renter@gmail.com", landlord_email),
    ("404", "78 Pine Road", "PROP-012", "s.hall.tenant@gmail.com", landlord_email),
    # 231 Elm Blvd
    ("1", "231 Elm Blvd", "PROP-013", "brian.w.apt@gmail.com", landlord_email),
    ("2", "231 Elm Blvd", "PROP-013", "a.allen.lease@gmail.com", landlord_email),
    ("3", "231 Elm Blvd", "PROP-013", "ryan.young.tenant@gmail.com", landlord_email),
    ("4", "231 Elm Blvd", "PROP-013", "nicole.king.apt@gmail.com", landlord_email),
    # 12 Oak Street
    ("2A", "12 Oak Street", "PROP-014", "j.wright.renter@gmail.com", landlord_email),
    ("2B", "12 Oak Street", "PROP-014", "megan.scott.unit@gmail.com", landlord_email),
    ("3A", "12 Oak Street", "PROP-014", "tyler.green.apt@gmail.com", landlord_email),
    ("3B", "12 Oak Street", "PROP-014", "kayla.adams.tenant@gmail.com", landlord_email),
    # 103 Birch Court
    ("201", "103 Birch Court", "PROP-015", "n.baker.renter@gmail.com", landlord_email),
    ("202", "103 Birch Court", "PROP-015", "b.carter.lease@gmail.com", landlord_email),
    ("301", "103 Birch Court", "PROP-015", "andrew.m.tenant@gmail.com", landlord_email),
    ("302", "103 Birch Court", "PROP-015", "s.perez.apt@gmail.com", landlord_email),
    # 55 Walnut Drive
    ("1A", "55 Walnut Drive", "PROP-016", "justin.r.renter@gmail.com", landlord_email),
    ("1B", "55 Walnut Drive", "PROP-016", "l.turner.tenant@gmail.com", landlord_email),
    ("2A", "55 Walnut Drive", "PROP-016", "b.phillips.apt@gmail.com", landlord_email),
    ("2B", "55 Walnut Drive", "PROP-016", "h.campbell.lease@gmail.com", landlord_email),
    # 18 Spruce Way
    ("101", "18 Spruce Way", "PROP-017", "eric.parker.unit@gmail.com", landlord_email),
    ("102", "18 Spruce Way", "PROP-017", "amber.evans.tenant@gmail.com", landlord_email),
    ("201", "18 Spruce Way", "PROP-017", "jon.edwards.apt@gmail.com", landlord_email),
    ("202", "18 Spruce Way", "PROP-017", "rachel.c.renter@gmail.com", landlord_email),
    # 400 Willow Terrace
    ("4", "400 Willow Terrace", "PROP-018", "aaron.s.tenant@gmail.com", landlord_email),
    ("5", "400 Willow Terrace", "PROP-018", "d.sanchez.apt@gmail.com", landlord_email),
    ("6", "400 Willow Terrace", "PROP-018", "z.morris.renter@gmail.com", landlord_email),
    ("7", "400 Willow Terrace", "PROP-018", "melissa.r.tenant@gmail.com", landlord_email),
    # 7 Ash Place
    ("A1", "7 Ash Place", "PROP-019", "p.reed.apt@gmail.com", landlord_email),
    ("A2", "7 Ash Place", "PROP-019", "tiffany.cook.lease@gmail.com", landlord_email),
    ("B1", "7 Ash Place", "PROP-019", "g.morgan.tenant@gmail.com", landlord_email),
    ("B2", "7 Ash Place", "PROP-019", "crystal.bell.apt@gmail.com", landlord_email),
    # 92 Hawthorn Gardens
    ("101", "92 Hawthorn Gardens", "PROP-020", "sean.murphy.renter@gmail.com", landlord_email),
    ("201", "92 Hawthorn Gardens", "PROP-020", "v.rivera.tenant@gmail.com", landlord_email),
    ("301", "92 Hawthorn Gardens", "PROP-020", "derek.cooper.apt@gmail.com", landlord_email),
    ("401", "92 Hawthorn Gardens", "PROP-020", "m.richardson.lease@gmail.com", landlord_email),
]

for i, u in enumerate(units, 2):
    add_row(ws_units, i, u)
auto_width(ws_units)

# =========================================================================
# TAB 5: Handymen
# =========================================================================
ws_handy = wb.create_sheet("Handymen")
headers = ["First Name", "Last Name", "Email", "Phone Number", "Description", "Location"]
style_sheet(ws_handy, headers)

handymen = [
    ("Mike", "Torres", "summit.elevator.svc@gmail.com", "(555) 234-8901", "Elevator repair & maintenance", "Metro area"),
    ("Bob", "Henderson", "bob.plumbing.co@gmail.com", "(555) 891-2345", "Plumbing — emergency & scheduled", "Metro area"),
    ("Tom", "Nguyen", "ace.hvac.quotes@gmail.com", "(555) 456-7890", "HVAC — heating & cooling systems", "Metro area"),
    ("Frank", "Russo", "cityelectric.bids@gmail.com", "(555) 567-1234", "Electrical — wiring, outlets, panels", "Metro area"),
    ("Dave", "Kowalski", "profix.contractors@gmail.com", "(555) 567-8901", "General contracting & repairs", "Metro area"),
    ("Steve", "Palmer", "quickrepair.quotes@gmail.com", "(555) 678-2345", "General repairs — appliances, fixtures", "Metro area"),
    ("Jim", "Kowalski", "metro.roofing.bids@gmail.com", "(555) 789-3456", "Roofing — repair & replacement", "Metro area"),
    ("Ray", "Thompson", "allpro.maintenance@gmail.com", "(555) 890-4567", "Property maintenance — monthly contract", "Metro area"),
    ("Lisa", "Tran", "sunrise.pest.quotes@gmail.com", "(555) 901-5678", "Pest control — residential & commercial", "Metro area"),
    ("Mark", "Chen", "premier.appliance.svc@gmail.com", "(555) 345-6789", "Appliance repair — all major brands", "Metro area"),
    ("Carlos", "Rivera", "greenleaf.quotes@gmail.com", "(555) 456-8901", "Landscaping — mowing, trimming, planting", "Metro area"),
    ("Tony", "Marchetti", "fastfix.gc.quotes@gmail.com", "(555) 567-9012", "General contracting — renovations", "Metro area"),
    ("Paul", "Anderson", "bluesky.wd.bids@gmail.com", "(555) 678-0123", "Windows & doors — repair & installation", "Metro area"),
    ("Jeff", "Williams", "safeguard.locksmith@gmail.com", "(555) 789-0123", "Locksmith — electronic & traditional locks", "Metro area"),
    ("Dan", "Murphy", "clearview.glass@gmail.com", "(555) 234-5678", "Glass & mirror — replacement & repair", "Metro area"),
    ("Rick", "Santos", "truseal.waterproof@gmail.com", "(555) 345-7890", "Waterproofing — basements & foundations", "Metro area"),
    ("Maria", "Lopez", "brightstar.cleaning@gmail.com", "(555) 456-0123", "Cleaning — common areas & deep clean", "Metro area"),
    ("Kevin", "O'Brien", "redline.firesafety@gmail.com", "(555) 567-1234", "Fire safety — alarms, extinguishers, inspections", "Metro area"),
    ("Carla", "Diaz", "ecomold.remediation@gmail.com", "(555) 678-3456", "Mold remediation & prevention", "Metro area"),
    ("Sam", "Patel", "atlas.drywall@gmail.com", "(555) 789-4567", "Drywall & painting — repair & installation", "Metro area"),
]

for i, h in enumerate(handymen, 2):
    add_row(ws_handy, i, h)
auto_width(ws_handy)

# Save
output = "Elron_Simulation_Data.xlsx"
wb.save(output)
print(f"Created {output} with {len(tenants)} tenants, {len(properties)} properties, {len(units)} units, {len(handymen)} handymen")
