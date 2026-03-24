"""
Generate realistic attachment files for the inbox simulator.

Produces real PDF, XLSX, JPEG, and PNG files in memory (as bytes)
so they can be attached to emails as proper MIME parts.
"""

import io
import random
from datetime import datetime, timedelta


def generate_pdf_devis(filename: str, **kwargs) -> bytes:
    """Generate a simple PDF quote/invoice document."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Header
    pdf.set_font("Helvetica", "B", 16)
    company = kwargs.get("company", "Entreprise Exemple")
    pdf.cell(0, 10, company, ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, kwargs.get("address", "12 rue du Commerce, 75015 Paris"), ln=True)
    pdf.cell(0, 5, kwargs.get("phone", "01 23 45 67 89"), ln=True)
    pdf.cell(0, 5, kwargs.get("email_addr", "contact@entreprise.fr"), ln=True)
    pdf.ln(10)

    # Title
    doc_type = kwargs.get("doc_type", "DEVIS")
    ref = kwargs.get("ref", "DEV-2026-001")
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"{doc_type} N. {ref}", ln=True)
    pdf.ln(5)

    # Date and client
    pdf.set_font("Helvetica", "", 10)
    date_str = kwargs.get("date", "15/03/2026")
    pdf.cell(0, 5, f"Date : {date_str}", ln=True)
    client = kwargs.get("client", "Elron Gestion Locative")
    pdf.cell(0, 5, f"Client : {client}", ln=True)
    pdf.cell(0, 5, kwargs.get("client_address", "24 avenue de la Republique, 93100 Montreuil"), ln=True)
    pdf.ln(10)

    # Table header
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(90, 8, "Description", border=1, fill=True)
    pdf.cell(30, 8, "Quantite", border=1, fill=True, align="C")
    pdf.cell(35, 8, "Prix unit. HT", border=1, fill=True, align="C")
    pdf.cell(35, 8, "Total HT", border=1, fill=True, align="C")
    pdf.ln()

    # Items
    items = kwargs.get("items", [
        ("Deplacement et diagnostic", 1, 50.00),
        ("Fourniture - piece de rechange", 1, 85.00),
        ("Main d'oeuvre", 2, 45.00),
    ])

    pdf.set_font("Helvetica", "", 10)
    total_ht = 0
    for desc, qty, price in items:
        line_total = qty * price
        total_ht += line_total
        pdf.cell(90, 7, desc, border=1)
        pdf.cell(30, 7, str(qty), border=1, align="C")
        pdf.cell(35, 7, f"{price:.2f} EUR", border=1, align="C")
        pdf.cell(35, 7, f"{line_total:.2f} EUR", border=1, align="C")
        pdf.ln()

    # Totals
    tva_rate = kwargs.get("tva_rate", 0.10)
    tva = total_ht * tva_rate
    total_ttc = total_ht + tva

    pdf.ln(5)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(120, 7, "")
    pdf.cell(35, 7, "Total HT :", align="R")
    pdf.cell(35, 7, f"{total_ht:.2f} EUR", align="R")
    pdf.ln()
    pdf.cell(120, 7, "")
    pdf.cell(35, 7, f"TVA ({int(tva_rate*100)}%) :", align="R")
    pdf.cell(35, 7, f"{tva:.2f} EUR", align="R")
    pdf.ln()
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(120, 7, "")
    pdf.cell(35, 7, "Total TTC :", align="R")
    pdf.cell(35, 7, f"{total_ttc:.2f} EUR", align="R")
    pdf.ln(15)

    # Footer
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5, "Conditions de paiement : 30 jours date de facture", ln=True)
    pdf.cell(0, 5, f"SIRET : {kwargs.get('siret', '123 456 789 00012')}", ln=True)

    return pdf.output()


def generate_pdf_report(filename: str, **kwargs) -> bytes:
    """Generate a simple PDF report (DPE, expertise, PV, etc.)."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()

    title = kwargs.get("title", "Rapport")
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, title, ln=True, align="C")
    pdf.ln(8)

    pdf.set_font("Helvetica", "", 10)
    date_str = kwargs.get("date", "15/03/2026")
    pdf.cell(0, 5, f"Date : {date_str}", ln=True)
    ref = kwargs.get("ref", "RPT-2026-001")
    pdf.cell(0, 5, f"Reference : {ref}", ln=True)
    address = kwargs.get("property_address", "")
    if address:
        pdf.cell(0, 5, f"Bien concerne : {address}", ln=True)
    pdf.ln(8)

    # Content paragraphs
    content = kwargs.get("content", [
        "Ce document presente les resultats de l'analyse effectuee.",
        "Les conclusions sont detaillees ci-dessous.",
    ])
    for para in content:
        pdf.multi_cell(0, 5, para)
        pdf.ln(3)

    # Optional table
    table_data = kwargs.get("table_data", None)
    if table_data:
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 10)
        for header in table_data["headers"]:
            pdf.cell(45, 7, header, border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 10)
        for row in table_data["rows"]:
            for cell in row:
                pdf.cell(45, 7, str(cell), border=1, align="C")
            pdf.ln()

    # Signature
    pdf.ln(15)
    pdf.set_font("Helvetica", "I", 10)
    author = kwargs.get("author", "Le responsable")
    pdf.cell(0, 5, author, ln=True, align="R")

    return pdf.output()


def generate_pdf_edl(filename: str, **kwargs) -> bytes:
    """Generate a PDF for etat des lieux (property inspection)."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    edl_type = kwargs.get("edl_type", "SORTIE")
    pdf.cell(0, 10, f"ETAT DES LIEUX DE {edl_type}", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Date : {kwargs.get('date', '20/03/2026')}", ln=True)
    pdf.cell(0, 5, f"Bien : {kwargs.get('address', '5 impasse des Acacias, Apt 1C')}", ln=True)
    pdf.cell(0, 5, f"Locataire : {kwargs.get('tenant', 'M. Rodriguez')}", ln=True)
    pdf.cell(0, 5, f"Representant bailleur : {kwargs.get('agent', 'Thomas Lefevre - Elron Gestion')}", ln=True)
    pdf.ln(8)

    # Room-by-room inspection
    rooms = kwargs.get("rooms", [
        ("Entree", "Bon etat", "Traces d'usure normales sur le sol"),
        ("Salon", "Correct", "4 trous non rebouches dans les murs"),
        ("Cuisine", "Degradation", "Brulure cigarette sur plan de travail"),
        ("Chambre", "Correct", "Tache sur moquette (30x20cm)"),
        ("Salle de bain", "Correct", "Joint silicone noirci a remplacer"),
        ("WC", "Bon etat", "RAS"),
    ])

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(40, 7, "Piece", border=1, fill=True)
    pdf.cell(40, 7, "Etat", border=1, fill=True, align="C")
    pdf.cell(110, 7, "Observations", border=1, fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    for room, state, obs in rooms:
        pdf.cell(40, 7, room, border=1)
        pdf.cell(40, 7, state, border=1, align="C")
        pdf.cell(110, 7, obs, border=1)
        pdf.ln()

    # Meters
    pdf.ln(8)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "Releves compteurs", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Eau : {kwargs.get('meter_water', '1247')} m3", ln=True)
    pdf.cell(0, 5, f"Electricite : {kwargs.get('meter_elec', '15823')} kWh", ln=True)

    # Keys
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "Cles remises", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, kwargs.get("keys", "2 cles porte entree, 1 cle boite aux lettres, 1 badge parking"), ln=True)

    # Signatures
    pdf.ln(15)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(95, 7, "Signature locataire :", border=0)
    pdf.cell(95, 7, "Signature bailleur :", border=0)
    pdf.ln(20)
    pdf.cell(95, 5, kwargs.get("tenant", "M. Rodriguez"), align="C")
    pdf.cell(95, 5, kwargs.get("agent", "Thomas Lefevre"), align="C")

    return pdf.output()


def generate_xlsx_comparatif(filename: str, **kwargs) -> bytes:
    """Generate an XLSX spreadsheet (comparatif, suivi charges, etc.)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = kwargs.get("sheet_title", "Comparatif")

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title row
    title = kwargs.get("title", "Comparatif EDL Entree / Sortie")
    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Date : {kwargs.get('date', '20/03/2026')}"
    ws["A3"] = f"Bien : {kwargs.get('address', '5 impasse des Acacias, Apt 1C')}"

    # Headers
    headers = kwargs.get("headers", ["Piece", "Etat entree", "Etat sortie", "Degradation", "Retenue (EUR)"])
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data
    rows = kwargs.get("rows", [
        ["Entree", "Bon etat", "Bon etat", "Non", 0],
        ["Salon", "Bon etat", "Trous murs", "Oui", 40],
        ["Cuisine", "Bon etat", "Brulure plan travail", "Oui", 150],
        ["Chambre", "Bon etat", "Tache moquette", "Oui", 80],
        ["Salle de bain", "Bon etat", "Joint noirci", "Oui", 30],
        ["WC", "Bon etat", "Bon etat", "Non", 0],
    ])

    for row_idx, row_data in enumerate(rows, 6):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Total row
    total_row = len(rows) + 6
    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(r[4] for r in rows if isinstance(r[4], (int, float)))).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_photo_jpg(filename: str, **kwargs) -> bytes:
    """Generate a simple JPEG image simulating a property photo."""
    from PIL import Image, ImageDraw, ImageFont

    width = kwargs.get("width", 800)
    height = kwargs.get("height", 600)
    bg_color = kwargs.get("bg_color", (240, 235, 220))

    img = Image.new("RGB", (width, height), bg_color)
    draw = ImageDraw.Draw(img)

    # Draw some rectangles to simulate room features
    rng = random.Random(filename)
    for _ in range(rng.randint(3, 8)):
        x1 = rng.randint(50, width - 200)
        y1 = rng.randint(50, height - 150)
        x2 = x1 + rng.randint(80, 250)
        y2 = y1 + rng.randint(60, 200)
        color = (rng.randint(180, 230), rng.randint(170, 220), rng.randint(160, 210))
        draw.rectangle([x1, y1, x2, y2], fill=color, outline=(150, 140, 130), width=2)

    # Add label
    label = kwargs.get("label", filename.replace(".jpg", "").replace("_", " "))
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 20)
    except (OSError, IOError):
        font = ImageFont.load_default()

    # Text background
    text_bbox = draw.textbbox((0, 0), label, font=font)
    tw = text_bbox[2] - text_bbox[0]
    th = text_bbox[3] - text_bbox[1]
    tx = (width - tw) // 2
    ty = height - 50
    draw.rectangle([tx - 5, ty - 5, tx + tw + 5, ty + th + 5], fill=(0, 0, 0, 180))
    draw.text((tx, ty), label, fill=(255, 255, 255), font=font)

    # Date stamp
    date_str = kwargs.get("date_stamp", "24/03/2026 09:15")
    try:
        small_font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 14)
    except (OSError, IOError):
        small_font = ImageFont.load_default()
    draw.text((width - 180, 15), date_str, fill=(255, 140, 0), font=small_font)

    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=75)
    return buf.getvalue()


def generate_photo_png(filename: str, **kwargs) -> bytes:
    """Generate a simple PNG image."""
    from PIL import Image, ImageDraw, ImageFont

    width = kwargs.get("width", 640)
    height = kwargs.get("height", 480)
    bg_color = kwargs.get("bg_color", (245, 245, 245))

    img = Image.new("RGB", (width, height), bg_color)
    draw = ImageDraw.Draw(img)

    label = kwargs.get("label", filename.replace(".png", "").replace("_", " "))
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 18)
    except (OSError, IOError):
        font = ImageFont.load_default()

    # Simple visual content
    rng = random.Random(filename)
    for _ in range(rng.randint(2, 5)):
        x1 = rng.randint(30, width - 150)
        y1 = rng.randint(30, height - 120)
        x2 = x1 + rng.randint(60, 200)
        y2 = y1 + rng.randint(50, 150)
        color = (rng.randint(180, 240), rng.randint(170, 230), rng.randint(160, 220))
        draw.rectangle([x1, y1, x2, y2], fill=color, outline=(130, 130, 130))

    draw.text((20, height - 35), label, fill=(80, 80, 80), font=font)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Registry: map attachment "type" to generator function
# ---------------------------------------------------------------------------

_GENERATORS = {
    "pdf_devis": generate_pdf_devis,
    "pdf_report": generate_pdf_report,
    "pdf_edl": generate_pdf_edl,
    "xlsx": generate_xlsx_comparatif,
    "jpg": generate_photo_jpg,
    "png": generate_photo_png,
}

# MIME types
_MIME_TYPES = {
    "pdf_devis": "application/pdf",
    "pdf_report": "application/pdf",
    "pdf_edl": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "jpg": "image/jpeg",
    "png": "image/png",
}

_EXTENSIONS = {
    "pdf_devis": ".pdf",
    "pdf_report": ".pdf",
    "pdf_edl": ".pdf",
    "xlsx": ".xlsx",
    "jpg": ".jpg",
    "png": ".png",
}


def generate_attachment(att_type: str, filename: str, **kwargs) -> tuple:
    """Generate an attachment and return (filename, mime_type, bytes_data).

    Parameters
    ----------
    att_type : str
        One of: pdf_devis, pdf_report, pdf_edl, xlsx, jpg, png
    filename : str
        The filename to use for the attachment
    **kwargs :
        Extra params passed to the generator function
    """
    gen = _GENERATORS.get(att_type)
    if not gen:
        raise ValueError(f"Unknown attachment type: {att_type}")

    data = gen(filename, **kwargs)
    mime = _MIME_TYPES[att_type]
    return filename, mime, data
